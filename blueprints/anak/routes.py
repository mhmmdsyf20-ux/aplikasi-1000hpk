"""
blueprints/anak/routes.py — Route CRUD data anak dan dashboard utama.

Routes:
    GET      /anak/dashboard          — Dashboard utama dengan summary cards
    GET      /anak/                   — Daftar anak dengan search & pagination
    GET/POST /anak/tambah             — Form tambah anak baru
    GET/POST /anak/<id>/edit          — Form edit data anak
    GET      /anak/<id>               — Detail anak + daftar imunisasi
"""

import csv
import io
from datetime import date, timedelta, datetime

from flask import render_template, redirect, url_for, flash, request
from flask_login import current_user

from blueprints.anak import anak_bp
from extensions import db
from models import Anak, Imunisasi
from services.anak_service import validate_anak_data
from services.master_service import master_only
from services.imunisasi_service import generate_jadwal_imunisasi, update_status_terlewat
_HAS_IMUNISASI_SERVICE = True


IMPORT_ANAK_HEADERS = [
    'nama',
    'tanggal_lahir',
    'jenis_kelamin',
    'nama_ibu',
    'no_hp_ortu',
    'alamat',
    'berat_lahir',
    'panjang_lahir',
]


def _normalize_import_cell(value):
    """Convert CSV/Excel cell values into strings used by existing validation."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _read_import_rows(file):
    """Return (fieldnames, [(row_index, row_dict), ...]) from CSV or Excel upload."""
    filename = file.filename.lower()

    if filename.endswith('.csv'):
        stream = io.TextIOWrapper(file.stream, encoding='utf-8-sig')
        reader = csv.DictReader(stream)
        rows = []
        for row_index, raw_row in enumerate(reader, start=2):
            row = {
                key.strip().lower(): _normalize_import_cell(value)
                for key, value in raw_row.items()
                if key
            }
            rows.append((row_index, row))
        return reader.fieldnames, rows

    if filename.endswith(('.xlsx', '.xlsm')):
        from openpyxl import load_workbook

        workbook = load_workbook(file.stream, read_only=True, data_only=True)
        sheet = workbook.active
        row_values = sheet.iter_rows(values_only=True)
        header_row = next(row_values, None)
        if not header_row:
            return None, []

        fieldnames = [_normalize_import_cell(value) for value in header_row]
        headers = [header.strip().lower() for header in fieldnames]
        rows = []
        for row_index, values in enumerate(row_values, start=2):
            row = {
                header: _normalize_import_cell(value)
                for header, value in zip(headers, values)
                if header
            }
            rows.append((row_index, row))
        return fieldnames, rows

    raise ValueError('unsupported_file_type')


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard
# ─────────────────────────────────────────────────────────────────────────────

@anak_bp.route('/dashboard')
@master_only
def dashboard():
    """Dashboard utama aplikasi 1000 HPK."""
    # Update status imunisasi terlewat setiap kali dashboard dibuka
    update_status_terlewat()

    today = date.today()
    next_30 = today + timedelta(days=30)

    total_anak = Anak.query.count()

    imunisasi_hari_ini = Imunisasi.query.filter(
        Imunisasi.tanggal_jadwal == today,
        Imunisasi.status == 'terjadwal',
    ).count()

    imunisasi_mendatang = Imunisasi.query.filter(
        Imunisasi.tanggal_jadwal > today,
        Imunisasi.tanggal_jadwal <= next_30,
        Imunisasi.status == 'terjadwal',
    ).count()

    imunisasi_terlewat = Imunisasi.query.filter(
        Imunisasi.status == 'terlewat',
    ).count()

    jadwal_mendatang = (
        Imunisasi.query
        .filter(
            Imunisasi.status.in_(['terjadwal', 'terlewat']),
        )
        .order_by(Imunisasi.tanggal_jadwal.asc())
        .limit(20)
        .all()
    )

    return render_template(
        'anak/dashboard.html',
        total_anak=total_anak,
        imunisasi_hari_ini=imunisasi_hari_ini,
        imunisasi_mendatang=imunisasi_mendatang,
        imunisasi_terlewat=imunisasi_terlewat,
        jadwal_mendatang=jadwal_mendatang,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Daftar Anak
# ─────────────────────────────────────────────────────────────────────────────

@anak_bp.route('/')
@master_only
def list_anak():
    """Daftar semua anak dengan fitur pencarian dan pagination."""
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 10

    query = Anak.query

    if q:
        like_q = f'%{q}%'
        query = query.filter(
            db.or_(
                Anak.nama.ilike(like_q),
                Anak.nama_ibu.ilike(like_q),
            )
        )

    pagination = query.order_by(Anak.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return render_template(
        'anak/list.html',
        anak_list=pagination.items,
        pagination=pagination,
        q=q,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Tambah Anak
# ─────────────────────────────────────────────────────────────────────────────

@anak_bp.route('/tambah', methods=['GET', 'POST'])
@master_only
def tambah_anak():
    """Form tambah anak baru."""
    if request.method == 'POST':
        data = {
            'nama': request.form.get('nama', '').strip(),
            'tanggal_lahir': request.form.get('tanggal_lahir', '').strip(),
            'jenis_kelamin': request.form.get('jenis_kelamin', '').strip(),
            'nama_ibu': request.form.get('nama_ibu', '').strip(),
            'no_hp_ortu': request.form.get('no_hp_ortu', '').strip(),
            'alamat': request.form.get('alamat', '').strip(),
            'berat_lahir': request.form.get('berat_lahir', '').strip(),
            'panjang_lahir': request.form.get('panjang_lahir', '').strip(),
        }

        errors = validate_anak_data(data)

        tanggal_lahir = None
        if data['tanggal_lahir']:
            try:
                tanggal_lahir = datetime.strptime(data['tanggal_lahir'], '%Y-%m-%d').date()
                # Validasi umur 0–2 tahun
                umur_hari = (date.today() - tanggal_lahir).days
                if umur_hari < 0:
                    errors.append('Tanggal lahir tidak boleh di masa depan.')
                elif umur_hari > 730:
                    errors.append('Aplikasi hanya untuk anak usia 0–2 tahun (maksimal 730 hari).')
            except ValueError:
                errors.append('Format tanggal lahir tidak valid.')

        if errors:
            for err in errors:
                flash(err, 'danger')
            return render_template('anak/form.html', data=data, mode='tambah',
                                   today=date.today().strftime('%Y-%m-%d'),
                                   min_date=(date.today() - timedelta(days=730)).strftime('%Y-%m-%d'))

        anak = Anak(
            nama=data['nama'],
            tanggal_lahir=tanggal_lahir,
            jenis_kelamin=data['jenis_kelamin'],
            nama_ibu=data['nama_ibu'],
            no_hp_ortu=data['no_hp_ortu'],
            alamat=data['alamat'] or None,
            berat_lahir=float(data['berat_lahir']) if data['berat_lahir'] else None,
            panjang_lahir=float(data['panjang_lahir']) if data['panjang_lahir'] else None,
            created_by=current_user.id,
        )
        db.session.add(anak)
        db.session.flush()

        try:
            jadwal_list = generate_jadwal_imunisasi(tanggal_lahir)
            for jadwal in jadwal_list:
                imun = Imunisasi(
                    anak_id=anak.id,
                    nama_vaksin=jadwal['nama_vaksin'],
                    tanggal_jadwal=jadwal['tanggal_jadwal'],
                    status='terjadwal',
                )
                db.session.add(imun)
        except Exception as e:
            flash(f'Peringatan: Gagal membuat jadwal imunisasi otomatis. {e}', 'warning')

        db.session.commit()
        flash(f'Data anak "{anak.nama}" berhasil disimpan.', 'success')
        return redirect(url_for('anak.detail_anak', anak_id=anak.id))

    return render_template('anak/form.html', data={}, mode='tambah',
                           today=date.today().strftime('%Y-%m-%d'),
                           min_date=(date.today() - timedelta(days=730)).strftime('%Y-%m-%d'))


# ─────────────────────────────────────────────────────────────────────────────
# Edit Anak
# ─────────────────────────────────────────────────────────────────────────────

@anak_bp.route('/<int:anak_id>/edit', methods=['GET', 'POST'])
@master_only
def edit_anak(anak_id):
    """Form edit data anak yang sudah ada."""
    anak = Anak.query.get_or_404(anak_id)

    if request.method == 'POST':
        data = {
            'nama': request.form.get('nama', '').strip(),
            'tanggal_lahir': request.form.get('tanggal_lahir', '').strip(),
            'jenis_kelamin': request.form.get('jenis_kelamin', '').strip(),
            'nama_ibu': request.form.get('nama_ibu', '').strip(),
            'no_hp_ortu': request.form.get('no_hp_ortu', '').strip(),
            'alamat': request.form.get('alamat', '').strip(),
            'berat_lahir': request.form.get('berat_lahir', '').strip(),
            'panjang_lahir': request.form.get('panjang_lahir', '').strip(),
        }

        errors = validate_anak_data(data)

        tanggal_lahir = None
        if data['tanggal_lahir']:
            try:
                tanggal_lahir = datetime.strptime(data['tanggal_lahir'], '%Y-%m-%d').date()
            except ValueError:
                errors.append('Format tanggal lahir tidak valid.')

        if errors:
            for err in errors:
                flash(err, 'danger')
            return render_template('anak/form.html', data=data, anak=anak, mode='edit',
                                   today=date.today().strftime('%Y-%m-%d'),
                                   min_date=(date.today() - timedelta(days=730)).strftime('%Y-%m-%d'))

        anak.nama = data['nama']
        anak.tanggal_lahir = tanggal_lahir
        anak.jenis_kelamin = data['jenis_kelamin']
        anak.nama_ibu = data['nama_ibu']
        anak.no_hp_ortu = data['no_hp_ortu']
        anak.alamat = data['alamat'] or None
        anak.berat_lahir = float(data['berat_lahir']) if data['berat_lahir'] else None
        anak.panjang_lahir = float(data['panjang_lahir']) if data['panjang_lahir'] else None

        db.session.commit()
        flash(f'Data anak "{anak.nama}" berhasil diperbarui.', 'success')
        return redirect(url_for('anak.detail_anak', anak_id=anak.id))

    data = {
        'nama': anak.nama,
        'tanggal_lahir': anak.tanggal_lahir.strftime('%Y-%m-%d') if anak.tanggal_lahir else '',
        'jenis_kelamin': anak.jenis_kelamin,
        'nama_ibu': anak.nama_ibu,
        'no_hp_ortu': anak.no_hp_ortu,
        'alamat': anak.alamat or '',
        'berat_lahir': str(anak.berat_lahir) if anak.berat_lahir else '',
        'panjang_lahir': str(anak.panjang_lahir) if anak.panjang_lahir else '',
    }

    return render_template('anak/form.html', data=data, anak=anak, mode='edit',
                           today=date.today().strftime('%Y-%m-%d'),
                           min_date=(date.today() - timedelta(days=730)).strftime('%Y-%m-%d'))


# ─────────────────────────────────────────────────────────────────────────────
# Detail Anak
# ─────────────────────────────────────────────────────────────────────────────

@anak_bp.route('/<int:anak_id>')
@master_only
def detail_anak(anak_id):
    """Halaman detail anak beserta daftar imunisasi."""
    anak = Anak.query.get_or_404(anak_id)
    imunisasi_list = (
        Imunisasi.query
        .filter_by(anak_id=anak_id)
        .order_by(Imunisasi.tanggal_jadwal.asc())
        .all()
    )

    return render_template(
        'anak/detail.html',
        anak=anak,
        imunisasi_list=imunisasi_list,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Import Data Anak
# ─────────────────────────────────────────────────────────────────────────────

@anak_bp.route('/import', methods=['GET', 'POST'])
@master_only
def import_anak():
    """Impor data anak dari file CSV atau Excel untuk mempercepat input data petugas."""
    success_count = 0
    row_errors = []
    template_headers = IMPORT_ANAK_HEADERS
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('File CSV atau Excel wajib dipilih.', 'danger')
            return render_template('anak/import.html', template_headers=template_headers)

        filename = file.filename.lower()
        if filename.endswith('.xls'):
            flash('File Excel lama .xls belum didukung. Simpan ulang sebagai .xlsx, lalu impor kembali.', 'danger')
            return render_template('anak/import.html', template_headers=template_headers)

        try:
            fieldnames, rows = _read_import_rows(file)
        except ValueError:
            flash('Hanya file CSV atau Excel .xlsx/.xlsm yang diizinkan.', 'danger')
            return render_template('anak/import.html', template_headers=template_headers)
        except Exception:
            flash('Gagal membaca file. Pastikan file CSV/Excel tidak rusak.', 'danger')
            return render_template('anak/import.html', template_headers=template_headers)

        if not fieldnames:
            flash('File kosong atau header tidak terdeteksi.', 'danger')
            return render_template('anak/import.html', template_headers=template_headers)

        headers = [h.strip().lower() for h in fieldnames if h]
        missing = [h for h in template_headers if h not in headers]
        if missing:
            flash(
                'Header file tidak lengkap. Pastikan header berikut ada: ' + ', '.join(missing),
                'danger'
            )
            return render_template('anak/import.html', template_headers=template_headers)

        for row_index, row in rows:
            if not any(row.values()):
                continue

            data = {
                'nama': row.get('nama', ''),
                'tanggal_lahir': row.get('tanggal_lahir', ''),
                'jenis_kelamin': row.get('jenis_kelamin', ''),
                'nama_ibu': row.get('nama_ibu', ''),
                'no_hp_ortu': row.get('no_hp_ortu', ''),
                'alamat': row.get('alamat', ''),
                'berat_lahir': row.get('berat_lahir', ''),
                'panjang_lahir': row.get('panjang_lahir', ''),
            }

            errors = validate_anak_data(data)
            tanggal_lahir = None
            if data['tanggal_lahir']:
                try:
                    tanggal_lahir = datetime.strptime(data['tanggal_lahir'], '%Y-%m-%d').date()
                    umur_hari = (date.today() - tanggal_lahir).days
                    if umur_hari < 0:
                        errors.append('Tanggal lahir tidak boleh di masa depan.')
                    elif umur_hari > 730:
                        errors.append('Aplikasi hanya untuk anak usia 0–2 tahun (maksimal 730 hari).')
                except ValueError:
                    errors.append('Format tanggal lahir tidak valid. Gunakan YYYY-MM-DD.')

            if data['jenis_kelamin'] not in ('L', 'P'):
                errors.append('Jenis kelamin harus L atau P.')

            try:
                berat_lahir = float(data['berat_lahir']) if data['berat_lahir'] else None
            except ValueError:
                errors.append('Berat lahir harus berupa angka.')
                berat_lahir = None

            try:
                panjang_lahir = float(data['panjang_lahir']) if data['panjang_lahir'] else None
            except ValueError:
                errors.append('Panjang lahir harus berupa angka.')
                panjang_lahir = None

            if errors:
                row_errors.append(f'Baris {row_index}: ' + '; '.join(errors))
                continue

            anak = Anak(
                nama=data['nama'],
                tanggal_lahir=tanggal_lahir,
                jenis_kelamin=data['jenis_kelamin'],
                nama_ibu=data['nama_ibu'],
                no_hp_ortu=data['no_hp_ortu'],
                alamat=data['alamat'] or None,
                berat_lahir=berat_lahir,
                panjang_lahir=panjang_lahir,
                created_by=current_user.id,
            )
            db.session.add(anak)
            db.session.flush()

            try:
                jadwal_list = generate_jadwal_imunisasi(tanggal_lahir)
                for jadwal in jadwal_list:
                    imun = Imunisasi(
                        anak_id=anak.id,
                        nama_vaksin=jadwal['nama_vaksin'],
                        tanggal_jadwal=jadwal['tanggal_jadwal'],
                        status='terjadwal',
                    )
                    db.session.add(imun)
            except Exception as e:
                row_errors.append(
                    f'Baris {row_index}: berhasil menyimpan anak, tetapi gagal membuat jadwal imunisasi otomatis ({e}).'
                )

            success_count += 1

        if success_count > 0:
            db.session.commit()

        if success_count:
            flash(f'{success_count} data anak berhasil diimpor.', 'success')
        if row_errors:
            for error in row_errors:
                flash(error, 'warning')

    return render_template('anak/import.html', template_headers=template_headers)


# ─────────────────────────────────────────────────────────────────────────────
# API Endpoints untuk Chart.js
# ─────────────────────────────────────────────────────────────────────────────

from flask import jsonify
from datetime import date as _date
from calendar import month_abbr


@anak_bp.route('/api/chart/status')
@master_only
def api_chart_status():
    """Endpoint JSON untuk donut chart status imunisasi."""
    selesai = Imunisasi.query.filter_by(status='selesai').count()
    terjadwal = Imunisasi.query.filter_by(status='terjadwal').count()
    terlewat = Imunisasi.query.filter_by(status='terlewat').count()
    return jsonify({"selesai": selesai, "terjadwal": terjadwal, "terlewat": terlewat})


@anak_bp.route('/api/chart/bulanan')
@master_only
def api_chart_bulanan():
    """Endpoint JSON untuk bar chart imunisasi selesai per bulan (6 bulan terakhir)."""
    from datetime import timedelta
    today = _date.today()
    labels = []
    values = []

    for i in range(5, -1, -1):
        # Hitung bulan ke-i yang lalu
        month = today.month - i
        year = today.year
        while month <= 0:
            month += 12
            year -= 1

        label = f"{month_abbr[month]} {year}"
        labels.append(label)

        # Hitung imunisasi selesai di bulan tersebut
        count = Imunisasi.query.filter(
            db.extract('year', Imunisasi.tanggal_realisasi) == year,
            db.extract('month', Imunisasi.tanggal_realisasi) == month,
            Imunisasi.status == 'selesai',
        ).count()
        values.append(count)

    return jsonify({"labels": labels, "values": values})


@anak_bp.route('/api/chart/progress/<int:anak_id>')
@master_only
def api_chart_progress(anak_id):
    """Endpoint JSON untuk progress imunisasi per anak."""
    from services.laporan_service import hitung_progress_anak
    progress = hitung_progress_anak(anak_id)
    total = Imunisasi.query.filter_by(anak_id=anak_id).count()
    selesai = Imunisasi.query.filter_by(anak_id=anak_id, status='selesai').count()
    return jsonify({"progress": progress, "selesai": selesai, "total": total})
