"""
blueprints/laporan/routes.py — Route laporan dan ekspor data (Admin only).

Routes:
    GET /laporan/              — Halaman laporan dengan filter tanggal
    GET /laporan/export/pdf    — Download laporan PDF
    GET /laporan/export/excel  — Download laporan Excel (.xlsx)
"""

import io
from datetime import date, datetime

from flask import render_template, request, flash, redirect, url_for, send_file, current_app
from flask_login import current_user

from blueprints.laporan import laporan_bp
from services.auth_service import role_required
from services.master_service import master_only
from services.laporan_service import get_laporan, get_statistik, get_cakupan_per_vaksin, hitung_progress_anak
from models import Anak


@laporan_bp.route("/")
@master_only
@role_required("admin")
def index():
    """Halaman laporan dengan filter tanggal dan statistik."""
    # Default: bulan ini
    today = date.today()
    default_start = today.replace(day=1).strftime('%Y-%m-%d')
    default_end = today.strftime('%Y-%m-%d')

    start_str = request.args.get("start_date", default_start)
    end_str = request.args.get("end_date", default_end)

    try:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    except ValueError:
        flash("Format tanggal tidak valid.", "danger")
        start_date = today.replace(day=1)
        end_date = today

    if start_date > end_date:
        flash("Tanggal mulai tidak boleh lebih besar dari tanggal akhir.", "danger")
        start_date, end_date = end_date, start_date

    laporan_data = get_laporan(start_date, end_date)
    statistik = get_statistik()
    cakupan_vaksin = get_cakupan_per_vaksin()

    from models import User
    petugas_map = {u.id: u.nama_lengkap for u in User.query.all()}

    return render_template(
        "laporan/index.html",
        laporan_data=laporan_data,
        statistik=statistik,
        cakupan_vaksin=cakupan_vaksin,
        start_date=start_date,
        end_date=end_date,
        petugas_map=petugas_map,
    )


@laporan_bp.route("/export/excel")
@master_only
@role_required("admin")
def export_excel():
    """Download laporan dalam format Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    start_str = request.args.get("start_date", date.today().replace(day=1).strftime('%Y-%m-%d'))
    end_str = request.args.get("end_date", date.today().strftime('%Y-%m-%d'))

    try:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    except ValueError:
        start_date = date.today().replace(day=1)
        end_date = date.today()

    laporan_data = get_laporan(start_date, end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Imunisasi"

    # Header
    headers = ["No", "Nama Anak", "Tanggal Lahir", "Umur (Bulan)", "Nama Vaksin",
               "Tanggal Jadwal", "Tanggal Realisasi", "Status", "Nama Petugas"]
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    from models import User
    petugas_map = {u.id: u.nama_lengkap for u in User.query.all()}

    for row_idx, imun in enumerate(laporan_data, 2):
        anak = imun.anak
        petugas_nama = petugas_map.get(imun.petugas_id, "—") if imun.petugas_id else "—"
        ws.append([
            row_idx - 1,
            anak.nama,
            anak.tanggal_lahir.strftime('%d/%m/%Y') if anak.tanggal_lahir else "—",
            anak.umur_bulan,
            imun.nama_vaksin,
            imun.tanggal_jadwal.strftime('%d/%m/%Y') if imun.tanggal_jadwal else "—",
            imun.tanggal_realisasi.strftime('%d/%m/%Y') if imun.tanggal_realisasi else "—",
            imun.status.capitalize(),
            petugas_nama,
        ])

    # Auto-width kolom
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"laporan_imunisasi_{start_str}_{end_str}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@laporan_bp.route("/export/pdf")
@master_only
@role_required("admin")
def export_pdf():
    """Download laporan dalam format PDF menggunakan ReportLab."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    start_str = request.args.get("start_date", date.today().replace(day=1).strftime('%Y-%m-%d'))
    end_str = request.args.get("end_date", date.today().strftime('%Y-%m-%d'))

    try:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    except ValueError:
        start_date = date.today().replace(day=1)
        end_date = date.today()

    laporan_data = get_laporan(start_date, end_date)
    statistik = get_statistik()
    nama_fasilitas = current_app.config.get("NAMA_FASILITAS", "Puskesmas")

    # Ambil nama petugas via query langsung (model Imunisasi tidak punya relasi petugas)
    from models import User
    petugas_map = {u.id: u.nama_lengkap for u in User.query.all()}

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Judul
    elements.append(Paragraph(f"Laporan Imunisasi — {nama_fasilitas}", styles['Title']))
    elements.append(Paragraph(
        f"Periode: {start_date.strftime('%d %B %Y')} s/d {end_date.strftime('%d %B %Y')}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.5*cm))

    # Statistik ringkasan
    elements.append(Paragraph(
        f"Total Anak: {statistik['total_anak']} | "
        f"Selesai: {statistik['total_selesai']} | "
        f"Terlewat: {statistik['total_terlewat']} | "
        f"Cakupan: {statistik['persentase_cakupan']}%",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.5*cm))

    # Tabel data
    headers = ["No", "Nama Anak", "Tgl Lahir", "Umur", "Vaksin",
               "Tgl Jadwal", "Tgl Realisasi", "Status", "Petugas"]
    table_data = [headers]

    for idx, imun in enumerate(laporan_data, 1):
        anak = imun.anak
        petugas_nama = petugas_map.get(imun.petugas_id, "—") if imun.petugas_id else "—"
        table_data.append([
            str(idx),
            anak.nama,
            anak.tanggal_lahir.strftime('%d/%m/%Y') if anak.tanggal_lahir else "—",
            f"{anak.umur_bulan} bln",
            imun.nama_vaksin,
            imun.tanggal_jadwal.strftime('%d/%m/%Y') if imun.tanggal_jadwal else "—",
            imun.tanggal_realisasi.strftime('%d/%m/%Y') if imun.tanggal_realisasi else "—",
            imun.status.capitalize(),
            petugas_nama,
        ])

    if len(table_data) > 1:
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("Tidak ada data untuk periode yang dipilih.", styles['Normal']))

    doc.build(elements)
    output.seek(0)

    filename = f"laporan_imunisasi_{start_str}_{end_str}.pdf"
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@laporan_bp.route("/export/anak/<int:anak_id>/excel")
@master_only
def export_excel_anak(anak_id):
    """Download laporan imunisasi per anak dalam format Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from models import Anak, Imunisasi

    anak = Anak.query.get_or_404(anak_id)
    imunisasi_list = (
        Imunisasi.query
        .filter_by(anak_id=anak_id)
        .order_by(Imunisasi.tanggal_jadwal.asc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Imunisasi {anak.nama}"

    # Info anak
    ws.append(["Nama Anak", anak.nama])
    ws.append(["Tanggal Lahir", anak.tanggal_lahir.strftime('%d/%m/%Y')])
    ws.append(["Nama Ibu", anak.nama_ibu])
    ws.append(["Umur", f"{anak.umur_bulan} bulan"])
    ws.append([])

    # Header tabel
    headers = ["No", "Nama Vaksin", "Tanggal Jadwal", "Tanggal Realisasi", "Status", "Catatan"]
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for idx, imun in enumerate(imunisasi_list, 1):
        ws.append([
            idx,
            imun.nama_vaksin,
            imun.tanggal_jadwal.strftime('%d/%m/%Y'),
            imun.tanggal_realisasi.strftime('%d/%m/%Y') if imun.tanggal_realisasi else "—",
            imun.status.capitalize(),
            imun.catatan or "—",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"imunisasi_{anak.nama.replace(' ', '_')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@laporan_bp.route("/export/anak/<int:anak_id>/pdf")
@master_only
def export_pdf_anak(anak_id):
    """Download laporan imunisasi per anak dalam format PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    from models import Anak, Imunisasi

    anak = Anak.query.get_or_404(anak_id)
    imunisasi_list = (
        Imunisasi.query
        .filter_by(anak_id=anak_id)
        .order_by(Imunisasi.tanggal_jadwal.asc())
        .all()
    )
    nama_fasilitas = current_app.config.get("NAMA_FASILITAS", "Puskesmas")

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Laporan Imunisasi — {nama_fasilitas}", styles['Title']))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(f"Nama Anak: {anak.nama}", styles['Normal']))
    elements.append(Paragraph(f"Tanggal Lahir: {anak.tanggal_lahir.strftime('%d %B %Y')}", styles['Normal']))
    elements.append(Paragraph(f"Nama Ibu: {anak.nama_ibu}", styles['Normal']))
    elements.append(Paragraph(f"Umur: {anak.umur_bulan} bulan", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    selesai = sum(1 for i in imunisasi_list if i.status == 'selesai')
    elements.append(Paragraph(
        f"Total Vaksin: {len(imunisasi_list)} | Selesai: {selesai} | "
        f"Belum: {len(imunisasi_list) - selesai}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.5*cm))

    headers = ["No", "Vaksin", "Tgl Jadwal", "Tgl Realisasi", "Status"]
    table_data = [headers]
    for idx, imun in enumerate(imunisasi_list, 1):
        table_data.append([
            str(idx),
            imun.nama_vaksin,
            imun.tanggal_jadwal.strftime('%d/%m/%Y'),
            imun.tanggal_realisasi.strftime('%d/%m/%Y') if imun.tanggal_realisasi else "—",
            imun.status.capitalize(),
        ])

    table = Table(table_data, repeatRows=1, colWidths=[1*cm, 5*cm, 3*cm, 3*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    filename = f"imunisasi_{anak.nama.replace(' ', '_')}.pdf"
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )
